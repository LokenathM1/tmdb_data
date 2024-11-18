import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed

def get_poster_url(movie_title, api_key):
    url = "https://api.themoviedb.org/3/search/movie"
    params = {
        "api_key": api_key,
        "query": movie_title
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()  # Raises an HTTPError for bad responses
        data = response.json()
        if data['results']:
            poster_path = data['results'][0]['poster_path']
            if poster_path:
                return f"https://image.tmdb.org/t/p/w500{poster_path}"
    except requests.RequestException as e:
        print(f"Error fetching data for {movie_title}: {e}")
    return None

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook('movie_data.xlsx')
sheet = wb.active
if sheet.cell(row=1, column=5).value != "Poster URL":
    sheet.cell(row=1, column=5).value = "Poster URL"
api_key = "bcb4389dbe6debbbb654e45c96c161d1"

# Create a ThreadPoolExecutor to handle multiple requests concurrently
executor = ThreadPoolExecutor(max_workers=20)
futures = []

for row in range(2, sheet.max_row + 1):
    movie_title = sheet.cell(row=row, column=1).value
    release_status = sheet.cell(row=row, column=2).value
    popularity = sheet.cell(row=row, column=3).value
    vote_count = sheet.cell(row=row, column=4).value

    # Check conditions: status must be "Released", popularity > 20, and vote count > 400
    if (movie_title and release_status == "Released" and
        float(popularity) > 20 and int(vote_count) > 400):
        future = executor.submit(get_poster_url, movie_title, api_key)
        futures.append((row, future))

# Process the futures as they complete
for row, future in futures:
    poster_url = future.result()
    if poster_url:
        sheet.cell(row=row, column=5).value = poster_url
    else:
        print(f"No poster found for {sheet.cell(row=row, column=1).value}")

# Save the modified Excel file
wb.save('updated_movie_data.xlsx')
print("Updated movie data saved successfully.")